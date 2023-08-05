import os
import xlrd
import xlwt
import openpyxl
import json
from collections import OrderedDict

def xlsx2json_2(args):
    wb = xlrd.open_workbook(args)
    shArr = wb.sheet_names() #获取所有sheet名
    json_path = os.path.split(args)[0]
    for shname in shArr:
        convert_list = []
        sh = wb.sheet_by_name(shname) #通过sheet名获得对应sheet内容
        title = sh.row_values(1)  #获取第二行的内容作为title
        for rownum in range(2, sh.nrows): #从第三行开始就是数据了
            rowvalue = sh.row_values(rownum) #每次拿一行数据
            single = OrderedDict()
            for colnum in range(1, len(rowvalue)): #将这行数据按列读取放入字典中
                single[title[colnum]] = rowvalue[colnum]
            convert_list.append(single) #读完之后将字段放入列表中

        j = json.dumps(convert_list, ensure_ascii=False)
        if not os.path.exists(json_path + '/json/'):
            os.mkdir(json_path + '/json/')
        with open(json_path + '/json/' + shname + '.json', "w", encoding="utf-8") as f:
            f.write(j)
            f.close()


def readFromJson(path):
    with open(path, 'r') as f:
        jsonData = json.load(f)
    return jsonData

def jsonPath2xlsx(path):
    filePaths = os.listdir(path)
    filePaths = filePaths[1:]
    sheetArr = []
    excel = openpyxl.Workbook()
    for item in filePaths:
        sheet = excel.create_sheet(item[: 6], index=0)
        sheetArr.append(sheet)

    index = 0
    for item in filePaths:

        data = readFromJson(path + item)
        length = len(data)
        i, j = 0, 0
        line0 = data[0]
        for k, v in line0.items():
            j += 1
            sheetArr[index].cell(row=1, column=j, value=k)
        while i < length:
            eachLine = data[i]
            j = 0
            i += 1;
            for v in eachLine.values():
                j += 1
                sheetArr[index].cell(row=i + 1, column=j, value=v)
        index += 1
    excel.save(path + "merge.xlsx")

if __name__ == '__main__':
    # pass
    xlsx2json_2('/Users/star_xlliu/Documents/testResult.xlsx')
    # jsonPath2xlsx('/Users/star_xlliu/Documents/json/')