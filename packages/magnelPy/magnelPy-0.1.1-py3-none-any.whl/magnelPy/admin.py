#__author__ = "magnelPy"
#__email__= "ruben.vancoile@ugent.be"
#__date__= "2018-11-21"

####################
## MODULE IMPORTS ##
####################

import numpy as np
import pandas as pd
import os, shutil
import openpyxl
from openpyxl import load_workbook #resolves issue with writing data to the correct Excel-sheet
import time

##############
## FUNCTION ##
##############

## print pd.DataFrame to *.xlsx ##
##################################

def df_writeToExcel(data,name,sheet,SW_overwrite=True):
	#--------------------------------------------------------------------------------------------------------
	# Print single pd.DataFrame or list of pd.DataFrame to specified excelfile and sheet
	# Input:
	#	data : pd.DataFrame / list of pd.DataFrame (df/list)
	#	name : (str) path of excel file (can be relative to working dir)
	# 	sheet : name / list of names (str / list) for the excel-sheets
	#	SW_overwrite : (boolean) overwrite existing excel or not
	#--------------------------------------------------------------------------------------------------------

	# remove '.xlsx' from name if specified
	if name[-5:]=='.xlsx': name=name[:-5]

	# set all input to list
	if isinstance(data,pd.DataFrame): data=[data]; sheet=[sheet]

	# check if excel path exists - SW_overwrite to True if non-existent
	if not os.path.isfile(name+'.xlsx'): SW_overwrite=True

	# print data to excel
	dfList_writeToExcel(data,name,sheet,SW_overwrite)

## read multiple sheets from *.xlsx ##
######################################

def df_readFromExcel(name,sheet=None):
	#--------------------------------------------------------------------------------------------------------
	# Return Dict of pd.DataFrames with data *.xlsx worksheets, worksheet names as Dict keys
	# Input:
	#	name : (str) path of excel sheet (can be relative to working dir)
	# 	sheet : name / list of names (str / list) for the excel-sheets; if None : all sheets are parsed
	#--------------------------------------------------------------------------------------------------------

	# read full *.xlsx file (suboptimal)
	xlsx=pd.ExcelFile(name)

	# syntax formatting
	if not isinstance(sheet,list):
		if sheet is None:
			sheet = xlsx.sheet_names
		else: sheet=[sheet]

	# read worksheets and store in out-Dict
	out={}
	for worksheet in sheet:
		out[worksheet]=pd.read_excel(xlsx,worksheet,index_col=0)

	return out

## clean folders ##
###################

def RemoveFolderData(folder):

	for the_file in os.listdir(folder):
		file_path=os.path.join(folder, the_file)
		if os.path.isfile(file_path):
			os.unlink(file_path)


##################
## AUX FUNCTION ##
##################

def dfList_writeToExcel(mytotallist,name,sheet,SW_overwrite=True):
	#--------------------------------------------------------------------------------------------------------
	# Print list of pd.DataFrame to specified excelfile and sheet
	# Input:
	#	mytotallist : (list) list of pd.DataFrames
	#	name : (str) path of excel sheet (can be relative to working dir)
	# 	sheet : (list) list of names (str) for the excel-sheets
	# Note1: overwrites existing excelfile and will throw an error when an existing file is open
	# Note2: renamed function Print_DataFrame() (rvcpy)
	# Known issue: duplicate name will result in additional sheets printed
	#--------------------------------------------------------------------------------------------------------
	writer=pd.ExcelWriter(name+'.xlsx',engine='openpyxl')
	if not SW_overwrite: writer.book=load_workbook(name+'.xlsx')
	for n,df in enumerate(mytotallist):
		Single_Print_DataFrame(df,writer,sheet[n])
	writer.save()
	writer.close()

def Single_Print_DataFrame(df,writer,sheet):
	#--------------------------------------------------------------------------------------------------------
	# Print single pd.DataFrame to specified excelfile and sheet
	# Input:
	#	df : pd.DataFrame
	#	writer : pd.ExcelWriter object
	# 	sheet : (str) name of excel-sheet
	# Note: overwrites existing excelfile and will throw an error when an existing file is open
	#--------------------------------------------------------------------------------------------------------
	df.to_excel(writer,sheet_name=sheet)

def TimeStamp():
	""" Provides unique timestamp

	Returns
	-------
	timestamp : str with unique timestamp
	 
	"""
	return str(int(time.time()*10**6))

#####################
## LEGACY FUNCTION ##
#####################

def writeToExcel(df,filename,sheetname=None):
	#--------------------------------------------------------------------------------------------------------
	# Print list of pd.DataFrame to specified excelfile and sheet
	# Input:
	#	df : pd.DataFrame
	#	name : (str) path of excel sheet (can be relative to working dir)
	# 	sheet : (str) name of excel-sheet
	# Note1: adds to existing excelfile and will throw an error when file does not exist
	# acknowledgement: TTH 11.2018 - function in GenerateRandom.py
	#--------------------------------------------------------------------------------------------------------
    writer = pd.ExcelWriter(filename+'.xlsx',engine='openpyxl')
    writer.book = load_workbook(filename+'.xlsx')
    if sheetname is not None: Single_Print_DataFrame(df,writer,sheetname)
    Single_Print_DataFrame(df,writer,'Sheet1')
    writer.save()

def Print_DataFrame(mytotallist,name,sheet,SW_overwrite=True):
	#--------------------------------------------------------------------------------------------------------
	# Print list of pd.DataFrame to specified excelfile and sheet
	# Input:
	#	mytotallist : (list) list of pd.DataFrames
	#	name : (str) path of excel sheet (can be relative to working dir)
	# 	sheet : (list) list of names (str) for the excel-sheets
	# Note1: overwrites existing excelfile and will throw an error when an existing file is open
	# Note2: renamed function Print_DataFrame() (rvcpy)
	#--------------------------------------------------------------------------------------------------------
	writer=pd.ExcelWriter(name+'.xlsx')
	for n,df in enumerate(mytotallist):
		Single_Print_DataFrame(df,writer,sheet[n])
	writer.save()
	writer.close()

#########################
## STAND ALONE - DEBUG ##
#########################

if __name__ == "__main__":

	## df_writeToExcel test ##
	if True:
		data=pd.DataFrame(0,index=['A','B'],columns=[1,2])
		name='test.xlsx'
		sheet='fourth'
		df_writeToExcel(data,name,sheet,False)


	## df_readFromExcel test ##
	if False:
		path="C:\\Users\\rvcoile\\Google Drive\\Research\\Codes\\Specific\\specific\\Nload_CDF - kopie.xlsx"
		sheet=['Iqbal_Fx','VanCoile_Fx']

		data=df_readFromExcel(path,sheet)

		print(data['VanCoile_Fx'])