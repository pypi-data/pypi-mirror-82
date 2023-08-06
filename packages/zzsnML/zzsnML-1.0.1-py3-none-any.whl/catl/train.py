# -*- coding: utf-8 -*-
"""
Created on Tue Jun 5 2018

@author: WuDaqing
"""

import os
import pickle   
from catl.utilities import preprocess_train
from catl.model import ensemble
from openpyxl import Workbook

name = input('Please input the name of company: ')
current_path = os.getcwd()
if os.path.isdir('data/'+name+'/preprocess') == False:
    os.makedirs(r'data/'+name+'/preprocess')
if os.path.isdir('results/'+name+'/train/model/') == False:
    os.makedirs(r'results/'+name+'/train/model/')
if os.path.isdir('results/'+name+'/train/results/') == False:
    os.makedirs(r'results/'+name+'/train/results/')
# print(os.getcwd())
preprocess = preprocess_train(name=name,path=r'data/'+name+'/'+name)
preprocess.read_excel()
Original_Data,Original_Data_Useless,Labels = preprocess.excel2sentences()

Vocabulary_Title = preprocess.get_vocabulary_title(title_weight=5,feature_ratio=0.1) # feature_ratio可调节，用来控制词表的长度，防止词表过长，运行时间太长或者内存溢出。
TFIDF_Title,IDF_Title = preprocess.get_tfidf_title(title_weight=5) # title_weight可调节，用于标题重复几次，增加标题的作用。
with open('data/'+name+'/preprocess/'+name+'_vocabulary_title.pkl','wb') as save1:
    pickle.dump(Vocabulary_Title,save1)
with open('data/'+name+'/preprocess/'+name+'_idf_title.pkl','wb') as save2:
    pickle.dump(IDF_Title,save2)

Model = ensemble(name=name,r=0.95,data=TFIDF_Title,labels=Labels,model_save_path='results/'+name+'/train/model/',results_save_path='results/'+name+'/train/results/') # r可调节，训练在召回率低于r时停止过滤进入下阶段过滤。
Threshold,Index_Retain_Predict_Title,Index_Delete_Title = Model.train_title()

Vocabulary_Content = preprocess.get_vocabulary_content(feature_ratio=0.2,index=Index_Retain_Predict_Title) # feature_ratio可调节，用来控制词表的长度，防止词表过长，运行时间太长或者内存溢出。
TFIDF_Content,IDF_Content = preprocess.get_tfidf_content(index=Index_Retain_Predict_Title)
with open('data/'+name+'/preprocess/'+name+'_vocabulary_content.pkl','wb') as save3:
    pickle.dump(Vocabulary_Content,save3)
with open('data/'+name+'/preprocess/'+name+'_idf_content.pkl','wb') as save4:
    pickle.dump(IDF_Content,save4)

threshold,Index_Retain_Predict_Content,Index_Delete_Content = Model.train_content(data=TFIDF_Content,r=0.9) # r可调节，训练最终在召回率低于r时终止。
with open('results/'+name+'/train/model/'+'title_threshold.pkl','wb') as save5:
    pickle.dump(Threshold,save5)
with open('results/'+name+'/train/model/'+'content_threshold.pkl','wb') as save6:
    pickle.dump(threshold,save6)

workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = 'finally'
worksheet1.cell(row=1,column=1).value = 'title'
worksheet1.cell(row=1,column=2).value = 'content'
worksheet1.cell(row=1,column=3).value = 'label'
for i in range(len(Index_Retain_Predict_Content)):
    worksheet1.cell(row=i+2,column=1).value = Original_Data[Index_Retain_Predict_Content[i]][0].encode('gbk','ignore').decode('gbk','ignore')
    worksheet1.cell(row=i+2,column=2).value = Original_Data[Index_Retain_Predict_Content[i]][1].encode('gbk','ignore').decode('gbk','ignore')
    worksheet1.cell(row=i+2,column=3).value = Original_Data[Index_Retain_Predict_Content[i]][2].encode('gbk','ignore').decode('gbk','ignore')
worksheet2 = workbook.create_sheet('delete through key words')
worksheet2.cell(row=1,column=1).value = 'title'
worksheet2.cell(row=1,column=2).value = 'content'
worksheet2.cell(row=1,column=3).value = 'label'
for i in range(len(Original_Data_Useless)):
    worksheet2.cell(row=i+2,column=1).value = Original_Data_Useless[i][0].encode('gbk','ignore').decode('gbk','ignore')
    worksheet2.cell(row=i+2,column=2).value = Original_Data_Useless[i][1].encode('gbk','ignore').decode('gbk','ignore')
    worksheet2.cell(row=i+2,column=3).value = Original_Data_Useless[i][2].encode('gbk','ignore').decode('gbk','ignore')
worksheet3 = workbook.create_sheet('delete through content')
worksheet3.cell(row=1,column=1).value = 'title'
worksheet3.cell(row=1,column=2).value = 'content'
worksheet3.cell(row=1,column=3).value = 'label'
for i in range(len(Index_Delete_Content)):
    worksheet3.cell(row=i+2,column=1).value = Original_Data[Index_Delete_Content[i]][0].encode('gbk','ignore').decode('gbk','ignore')
    worksheet3.cell(row=i+2,column=2).value = Original_Data[Index_Delete_Content[i]][1].encode('gbk','ignore').decode('gbk','ignore')
    worksheet3.cell(row=i+2,column=3).value = Original_Data[Index_Delete_Content[i]][2].encode('gbk','ignore').decode('gbk','ignore')
for ite in range(len(Index_Delete_Title)):
    worksheet = workbook.create_sheet('delete through title '+str(ite+1))
    worksheet.cell(row=1,column=1).value = 'title'
    worksheet.cell(row=1,column=2).value = 'content'
    worksheet.cell(row=1,column=3).value = 'label'   
    for i in range(len(Index_Delete_Title[ite+1])):
        worksheet.cell(row=i+2,column=1).value = Original_Data[Index_Delete_Title[ite+1][i]][0].encode('gbk','ignore').decode('gbk','ignore')
        worksheet.cell(row=i+2,column=2).value = Original_Data[Index_Delete_Title[ite+1][i]][1].encode('gbk','ignore').decode('gbk','ignore')
        worksheet.cell(row=i+2,column=3).value = Original_Data[Index_Delete_Title[ite+1][i]][2].encode('gbk','ignore').decode('gbk','ignore')
workbook.save('results/'+name+'/train/results/train_results.xlsx')