# -*- coding: utf-8 -*-
"""
Created on Tue Jun 5 2018

@author: WuDaqing
"""

import os
import pickle
import xlrd
import re
import jieba
from openpyxl import Workbook 
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.preprocessing import normalize
from sklearn import metrics
from openpyxl import Workbook
from sklearn.externals import joblib

class preprocess_train(object):
    def __init__(self,name,path):
        self.name = name
        self.path = path

        self.chinese_stopwords = []
        home_path = os.path.dirname(os.path.realpath(__file__))
        stopwords_path = os.path.join(home_path,'data/stopwords.txt')
        for line in open(stopwords_path,'rb'):
            self.chinese_stopwords.append(line.decode('utf-8-sig').split()[0])
        self.key_words = []
        for line in open(self.path+'_original_key_words.txt','rb'):
            self.key_words.append(line.decode('utf-8-sig').split()[0])
        jieba.load_userdict(self.path+'_original_key_words.txt')

    def read_excel(self):
        excel = xlrd.open_workbook(self.path+'_train.xls')
        table = excel.sheet_by_index(0)
        num_rows = table.nrows-1

        self.original_data = []
        for idx in range(1,num_rows+1):
            row = table.row_values(idx) 
            self.original_data.append(row)
        self.data = list(map(list,zip(*self.original_data)))
        self.labels = [int(self.data[2][i]=='保留') for i in range(num_rows)]

    def document2sentences(self,document):   
        symbols = frozenset(u"，。！？\n：；“”|）\u3000")  
        sentences= []
        tmp = []
        for character in document:    
            if not symbols.__contains__(character):
                tmp.append(character)
            elif character in "，。！？\n：；“”|）":
                tmp.append("。")
                for i in range(len(self.key_words)):
                    if self.key_words[i] in ''.join(tmp):
                        sentences.append(''.join(tmp))
                tmp = []
            elif character == "\u3000":
                continue
        for i in range(len(self.key_words)):
            if self.key_words[i] in ''.join(tmp):
                sentences.append(''.join(tmp))
        return ''.join(sentences)

    def filtrate_words(self,words):
        find_chinese = re.compile(u"[\u4e00-\u9fa5]+")
        symbols = "[A-Za-z0-9\[\`\~\!\@\#\$\^\&\*\(\)\=\|\{\}\'\:\;\'\,\[\]\.\<\>\/\?\~\！\@\#\\\&\*\%]"        
        filtrated_words = []
        for j in range(len(words)):
            if re.findall(find_chinese,words[j]) == []:
                continue
            elif re.sub(symbols, "",re.findall(find_chinese,words[j])[0]) == '':
                continue
            elif re.sub(symbols, "",re.findall(find_chinese,words[j])[0]) in self.chinese_stopwords:
                continue
            else:
                filtrated_words.append(re.sub(symbols, "",re.findall(find_chinese,words[j])[0]))    
        return ' '.join(filtrated_words)

    def excel2sentences(self):
        contents = self.data[1]
        print(self.name+' | Train | Content | Document 2 Sentences ......')
        contents_sentences = [self.document2sentences(document) for document in contents]
        original_data_useless = [self.original_data[i] for i in range(len(contents_sentences)) if contents_sentences[i] == '']
        self.original_data = [self.original_data[i] for i in range(len(contents_sentences)) if contents_sentences[i] != '']
        self.labels = [self.labels[i] for i in range(len(contents_sentences)) if contents_sentences[i] != '']
        titles = [self.data[0][i] for i in range(len(contents_sentences)) if contents_sentences[i] != '']
        contents_sentences = [contents_sentences[i] for i in range(len(contents_sentences)) if contents_sentences[i] != '']

        workbook = Workbook()
        worksheet1 = workbook.active
        worksheet1.title = 'use'
        worksheet1.cell(row=1,column=1).value = 'title'
        worksheet1.cell(row=1,column=2).value = 'content'
        worksheet1.cell(row=1,column=3).value = 'label'
        for i in range(len(self.original_data)):
            print(i)
            worksheet1.cell(row=i+2,column=1).value = self.original_data[i][0].encode('gbk','ignore').decode('gbk','ignore')
            worksheet1.cell(row=i+2,column=2).value = contents_sentences[i].encode('gbk','ignore').decode('gbk','ignore')
            worksheet1.cell(row=i+2,column=3).value = self.original_data[i][2].encode('gbk','ignore').decode('gbk','ignore')
        worksheet2 = workbook.create_sheet('useless')
        worksheet2.cell(row=1,column=1).value = 'title'
        worksheet2.cell(row=1,column=2).value = 'content'
        worksheet2.cell(row=1,column=3).value = 'label'
        for i in range(len(original_data_useless)):
            worksheet2.cell(row=i+2,column=1).value = original_data_useless[i][0].encode('gbk','ignore').decode('gbk','ignore')
            worksheet2.cell(row=i+2,column=2).value = original_data_useless[i][1].encode('gbk','ignore').decode('gbk','ignore')
            worksheet2.cell(row=i+2,column=3).value = original_data_useless[i][2].encode('gbk','ignore').decode('gbk','ignore')
        workbook.save(self.path+'_train_sentences.xlsx')

        print(self.name+' | Train |  Title  |  Tokenized ......')
        titles_tokenized = [jieba.lcut(sentences) for sentences in titles]
        print(self.name+' | Train | Content |  Tokenized ......')
        contents_sentences_tokenized = [jieba.lcut(sentences) for sentences in contents_sentences]
        print(self.name+' | Train |  Title  |  Filtered ......')
        self.titles_tokenized_filtered = [self.filtrate_words(words) for words in titles_tokenized]
        print(self.name+' | Train | Content |  Filtered ......')
        self.contents_sentences_tokenized_filtered = [self.filtrate_words(words) for words in contents_sentences_tokenized]

        return self.original_data,original_data_useless,self.labels
    
    def get_chi(self,data,labels):
        num = len(data)
        length = len(data[0])

        data_p = [data[i] for i in range(num) if labels[i]==1]
        data_n = [data[i] for i in range(num) if labels[i]==0]
        num_p = len(data_p)
        num_n = len(data_n)

        data_p_t = list(map(list,zip(*data_p)))
        data_n_t = list(map(list,zip(*data_n)))

        chi_square = []
        for i in range(length):
            b = data_p_t[i].count(0)
            d = data_n_t[i].count(0)
            a = num_p-b
            c = num_n-d
            if num_p*num_n*(a+c)*(b+d) == 0:
                chi_square.append(0)
            else:
                chi_square.append((num*pow(a*d-b*c,2))/(num_p*num_n*(a+c)*(b+d)))
        return chi_square

    def get_vocabulary_title(self,title_weight,feature_ratio):
        data = [title_weight*(self.titles_tokenized_filtered[i]+' ')+self.contents_sentences_tokenized_filtered[i] for i in range(len(self.labels))]
        labels = self.labels
        tf_transformer = CountVectorizer(ngram_range=(1,3))
        tf = tf_transformer.fit_transform(data)
        vocabulary_list = tf_transformer.get_feature_names()
        print(self.name+' | Train |  Title  | Vocabulary | Original Length | ' + str(len(vocabulary_list)))
        num_key_words = int(len(vocabulary_list)*feature_ratio)
        print(self.name+' | Train |  Title  | Vocabulary |     Length      | ' + str(num_key_words))
        tf_weights = tf.toarray().tolist()
        chi_square = self.get_chi(tf_weights,labels)

        print(self.name+' | Train |  Title  | Vocabulary | Complete by CHI ......')
        original_vocabulary_chi_square = [(vocabulary_list[i],chi_square[i]) for i in range(len(vocabulary_list))]
        sorted_original_vocabulary_chi_square = sorted(original_vocabulary_chi_square,key=lambda x:x[1],reverse=True)
        vocabulary_list = [sorted_original_vocabulary_chi_square[i][0] for i in range(num_key_words)]

        self.vocabulary_title = {}
        k = 0
        for word in vocabulary_list:
            self.vocabulary_title[word] = k
            k += 1
        return self.vocabulary_title
    def get_tfidf_title(self,title_weight):
        data = [title_weight*(self.titles_tokenized_filtered[i]+' ')+self.contents_sentences_tokenized_filtered[i] for i in range(len(self.labels))]
        tf_transformer = CountVectorizer(ngram_range=(1,3),vocabulary=self.vocabulary_title)
        train_tf = tf_transformer.fit_transform(data)
        print(self.name+' | Train |  Title  | TF | Completed ......') 
        tfidf_transformer = TfidfTransformer(norm='l2',use_idf=True,smooth_idf=True)
        train_tfidf = tfidf_transformer.fit_transform(train_tf)
        train_tfidf_weights = train_tfidf.toarray().tolist()
        print(self.name+' | Train |  Title  | TFIDF | Completed ......') 
        idf = tfidf_transformer.idf_.tolist()
        return train_tfidf_weights,idf 

    def get_vocabulary_content(self,feature_ratio,index):
        data = [self.contents_sentences_tokenized_filtered[idx] for idx in index]
        labels = [self.labels[idx] for idx in index]
        tf_transformer = CountVectorizer(ngram_range=(1,3))
        tf = tf_transformer.fit_transform(data)
        vocabulary_list = tf_transformer.get_feature_names()
        print(self.name+' | Train | Content | Vocabulary | Original Length | ' + str(len(vocabulary_list)))
        num_key_words = int(len(vocabulary_list)*feature_ratio)
        print(self.name+' | Train | Content | Vocabulary |     Length      | ' + str(num_key_words))
        tf_weights = tf.toarray().tolist()
        chi_square = self.get_chi(tf_weights,labels)

        print(self.name+' | Train | Content | Vocabulary | Complete by CHI ......')
        original_vocabulary_chi_square = [(vocabulary_list[i],chi_square[i]) for i in range(len(vocabulary_list))]
        sorted_original_vocabulary_chi_square = sorted(original_vocabulary_chi_square,key=lambda x:x[1],reverse=True)
        vocabulary_list = [sorted_original_vocabulary_chi_square[i][0] for i in range(num_key_words)]

        self.vocabulary_content = {}
        k = 0
        for word in vocabulary_list:
            self.vocabulary_content[word] = k
            k += 1
        return self.vocabulary_content
    def get_tfidf_content(self,index):
        data = [self.contents_sentences_tokenized_filtered[idx] for idx in index]
        tf_transformer = CountVectorizer(ngram_range=(1,3),vocabulary=self.vocabulary_content)
        train_tf = tf_transformer.fit_transform(data)
        print(self.name+' | Train | Content | TF | Completed ......') 
        tfidf_transformer = TfidfTransformer(norm='l2',use_idf=True,smooth_idf=True)
        train_tfidf = tfidf_transformer.fit_transform(train_tf)
        train_tfidf_weights = train_tfidf.toarray().tolist()
        print(self.name+' | Train | Content | TFIDF | Completed ......') 
        idf = tfidf_transformer.idf_.tolist()
        return train_tfidf_weights,idf

class single_predict(object):
    def __init__(self,name,title,content):
        self.name = name
        self.title = title
        self.content = content
        current_path = os.getcwd()
        if os.path.isdir('results/'+self.name+'/predict/results/') == False:
            os.makedirs(r'results/'+self.name+'/predict/results/')
        self.path = 'data/'+self.name+'/'+self.name
        self.model_load_path = 'results/'+self.name+'/train/model/'

        self.chinese_stopwords = []
        file_path = os.path.dirname(os.path.realpath(__file__))
        for line in open(os.path.join(file_path, 'data/stopwords.txt'),'rb'):
            self.chinese_stopwords.append(line.decode('utf-8-sig').split()[0])
        self.key_words = []
        for line in open(self.path+'_original_key_words.txt','rb'):
            self.key_words.append(line.decode('utf-8-sig').split()[0])
        jieba.load_userdict(self.path+'_original_key_words.txt')

        with open('data/'+self.name+'/preprocess/'+self.name+'_vocabulary_title.pkl','rb') as load1:
            self.vocabulary_title = pickle.load(load1)
        with open('data/'+self.name+'/preprocess/'+self.name+'_idf_title.pkl','rb') as load2:
            self.idf_title = pickle.load(load2)
        with open('results/'+self.name+'/train/model/'+'title_threshold.pkl','rb') as load3:
            self.Threshold = pickle.load(load3)
        with open('results/'+self.name+'/train/model/'+'content_threshold.pkl','rb') as load4:
            self.threshold = pickle.load(load4)
        with open('data/'+self.name+'/preprocess/'+self.name+'_vocabulary_content.pkl','rb') as load5:
            self.vocabulary_content = pickle.load(load5)
        with open('data/'+self.name+'/preprocess/'+self.name+'_idf_content.pkl','rb') as load6:
            self.idf_content = pickle.load(load6)

    def document2sentences(self,document):   
        symbols = frozenset(u"，。！？\n：；“”|）\u3000")  
        sentences= []
        tmp = []
        for character in document:    
            if not symbols.__contains__(character):
                tmp.append(character)
            elif character in "，。！？\n：；“”|）":
                tmp.append("。")
                for i in range(len(self.key_words)):
                    if self.key_words[i] in ''.join(tmp):
                        sentences.append(''.join(tmp))
                tmp = []
            elif character == "\u3000":
                continue
        for i in range(len(self.key_words)):
            if self.key_words[i] in ''.join(tmp):
                sentences.append(''.join(tmp))
        return ''.join(sentences)

    def filtrate_words(self,words):
        find_chinese = re.compile(u"[\u4e00-\u9fa5]+")
        symbols = "[A-Za-z0-9\[\`\~\!\@\#\$\^\&\*\(\)\=\|\{\}\'\:\;\'\,\[\]\.\<\>\/\?\~\！\@\#\\\&\*\%]"        
        filtrated_words = []
        for j in range(len(words)):
            if re.findall(find_chinese,words[j]) == []:
                continue
            elif re.sub(symbols, "",re.findall(find_chinese,words[j])[0]) == '':
                continue
            elif re.sub(symbols, "",re.findall(find_chinese,words[j])[0]) in self.chinese_stopwords:
                continue
            else:
                filtrated_words.append(re.sub(symbols, "",re.findall(find_chinese,words[j])[0]))    
        return ' '.join(filtrated_words)

    def predict(self):
        content_sentences = self.document2sentences(self.content)
        if content_sentences == '':
            prediction = '删除'
        else:
            title_tokenized = jieba.lcut(self.title)
            content_sentences_tokenized = jieba.lcut(content_sentences)
            title_tokenized_filtered = self.filtrate_words(title_tokenized)
            content_sentences_tokenized_filtered = self.filtrate_words(content_sentences_tokenized)
            data_title = [5*(title_tokenized_filtered+' ')+content_sentences_tokenized_filtered]
            tf_transformer_title = CountVectorizer(ngram_range=(1,3),vocabulary=self.vocabulary_title)
            tf_title = tf_transformer_title.fit_transform(data_title)
            tf_weight_title = tf_title.toarray().tolist()
            tfidf_weight_title = normalize([[x*y for x,y in zip(tf_weight_title[0],self.idf_title)]], norm='l2').tolist()
            for ite in range(1,len(self.Threshold)+1):
                clf_title = joblib.load(self.model_load_path+self.name+'_iteration_'+str(ite)+'_train_title_classifier.m')
                tmp = clf_title.predict_proba(tfidf_weight_title).tolist()
                if tmp[0][1] < self.Threshold[ite]:
                    prediction = '删除'
                    ite -= 1
                    break
                else:
                    continue
            if ite == len(self.Threshold):
                data_content = [content_sentences_tokenized_filtered]
                tf_transformer_content = CountVectorizer(ngram_range=(1,3),vocabulary=self.vocabulary_content)
                tf_content = tf_transformer_content.fit_transform(data_content)
                tf_weight_content = tf_content.toarray().tolist()
                tfidf_weight_content = normalize([[x*y for x,y in zip(tf_weight_content[0],self.idf_content)]], norm='l2').tolist()
                clf_content = joblib.load(self.model_load_path+self.name+'_train_content_classifier.m')
                tmp = clf_content.predict_proba(tfidf_weight_content).tolist()
                if tmp[0][1] < self.threshold:
                    prediction = '删除'
                else:
                    prediction = '保留'
        return prediction