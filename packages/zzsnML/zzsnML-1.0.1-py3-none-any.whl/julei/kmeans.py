# -*- coding: utf-8 -*-
"""
Created on Wed April 17 22:06:20 2019

@author: Wu-Daqing
"""

import time
import xlrd
import os
import math
import pickle
import numpy as np
from openpyxl import Workbook
from sklearn.cluster import KMeans
from julei.representation import Representation
class Kmeans:
    def __init__(self):
        pass
    def make_dir(self, path):
        dir_path = os.path.join(os.getcwd(), path)
        if os.path.isdir(dir_path) == False:
            os.makedirs(dir_path)

    def load_pkl(self):
        print(time.strftime('%Y-%m-%d %H:%M:%S'), '开始导入数据')
        representation = Representation()
        data = representation.text_represent()
        # with open('result/representation/text_representation.pkl','rb') as load1:
        #     data = pickle.load(load1)
        print(time.strftime('%Y-%m-%d %H:%M:%S'),'完成导入数据')
        num_data = data.shape[0]
        print("====================num_data = " + str(num_data))
        return data,num_data

    #tmp add:

    #print("---------------------num_class = "+str(num_class))
    def train(self, path ='result/kmeans/'):
        data,num_data = self.load_pkl()
        #num_class = 20
        num_class = int(math.sqrt(num_data))
        # print(num_class)
        print(time.strftime('%Y-%m-%d %H:%M:%S'),'开始训练模型')
        kmeans = KMeans(n_clusters=num_class, init='k-means++', n_init=5, max_iter=100)
        model = kmeans.fit(data)
        print(time.strftime('%Y-%m-%d %H:%M:%S'),'完成训练模型')
        classes = model.labels_
        centroids = model.cluster_centers_
        result = [[] for j in range(num_class)]
        data_cluster = [[] for j in range(num_class)]
        for i in range(num_data):
            for j in range(num_class):
                if classes[i] == j:
                    result[j].append(i)
                    data_cluster[j].append(data[i])
        print(time.strftime('%Y-%m-%d %H:%M:%S'),'完成计算结果')

        result_sorted = []
        similarity = []
        for j in range(num_class):
            distances = [(np.linalg.norm(centroids[j] - data_cluster[j][i]),result[j][i]) for i in range(len(result[j]))]
            distances_sorted = sorted(distances, key=lambda x: x[0])
            result_sorted.append([value[1] for value in distances_sorted])
            similarity.append([value[0] for value in distances_sorted])
        print(time.strftime('%Y-%m-%d %H:%M:%S'),'完成排序结果')
        
        with open(os.path.join(os.getcwd(), path)+'centroids.pkl','wb') as save1:
            pickle.dump(centroids,save1)
        return num_class,result_sorted,similarity

    def load_data(self, excel_path):
        excel = xlrd.open_workbook(excel_path)
        table = excel.sheet_by_index(0)
        num_rows = table.nrows-1
        content = []
        for idx in range(1,num_rows+1):
            row = table.row_values(idx)
            content.append([row[0],row[1],row[2],row[3],row[4],row[5],row[6]])#jjia 3ge
        return content

    def write(self, path ='result/kmeans/'):
        self.make_dir(path)
        file_names = sorted(os.listdir('data/'))
        original_data = []
        for file_name in file_names:
            if file_name[-5:] == '.xlsx':
                content = self.load_data(excel_path='data/'+file_name)
                print(time.strftime('%Y-%m-%d %H:%M:%S'),file_name.split('_')[0],'文本读取完毕')
                original_data += content

        print(time.strftime('%Y-%m-%d %H:%M:%S'),'开始写入结果')
        num_class,result_sorted,similarity = self.train(path)
        for j in range(num_class):
            print(time.strftime('%Y-%m-%d %H:%M:%S'),'第',j+1,'类有',len(result_sorted[j]),'条文本')
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = str(len(result_sorted[j]))
            worksheet.cell(row=1,column=1).value = '日期'
            worksheet.cell(row=1,column=2).value = '时间'
            worksheet.cell(row=1,column=3).value = '距离中心欧氏距离'
            worksheet.cell(row=1,column=4).value = '内容'
            worksheet.cell(row=1, column=5).value = '来源' #新加
            worksheet.cell(row=1, column=6).value = '标题'  # 新加
            worksheet.cell(row=1, column=7).value = '链接'  # 新加
            count = 1
            for i in range(len(result_sorted[j])):
                try:
                    worksheet.cell(row=count+1,column=5).value = original_data[result_sorted[j][i]][3].encode('gbk','ignore').decode('gbk','ignore')#新加
                    worksheet.cell(row=count + 1, column=6).value = original_data[result_sorted[j][i]][5].encode('gbk','ignore').decode( 'gbk', 'ignore')  # 新加

                    worksheet.cell(row=count + 1, column=7).value = original_data[result_sorted[j][i]][6].encode('gbk',
                                                                                                                 'ignore').decode(
                        'gbk', 'ignore')  # 新加
                   # print(original_data[result_sorted[j][i]][3])
                    worksheet.cell(row=count+1,column=4).value = original_data[result_sorted[j][i]][2].encode('gbk','ignore').decode('gbk', 'ignore')
                    worksheet.cell(row=count+1,column=1).value = original_data[result_sorted[j][i]][0].encode('gbk','ignore').decode('gbk','ignore')
                    worksheet.cell(row=count+1,column=2).value = original_data[result_sorted[j][i]][1].encode('gbk','ignore').decode('gbk','ignore')
                    worksheet.cell(row=count+1,column=3).value = similarity[j][i]
                    count += 1
                except Exception as e:
                    print('str(e):\t\t', str(e))
                    continue
            workbook.save(os.path.join(os.getcwd(), path)+str(count-1)+'_'+str(j+1)+'.xlsx')
            print(time.strftime('%Y-%m-%d %H:%M:%S'),j+1,'类写入Excel完毕','\n')
# km = Kmeans()
# km.write()