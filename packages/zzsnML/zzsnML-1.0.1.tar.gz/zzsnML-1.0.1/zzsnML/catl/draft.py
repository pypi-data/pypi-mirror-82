# -*- coding: utf-8 -*-
"""
Created on Tue Jun 5 2018

@author: WuDaqing
"""

import os
import pickle
import xlrd
import re
import jieba
from openpyxl import Workbook 
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.preprocessing import normalize
from sklearn import metrics
from sklearn.externals import joblib

def document2sentences(document,key_words):   
    symbols = frozenset(u"，。！？\n：；“”|）\u3000")  
    sentences= []
    tmp = []
    for character in document:    
        if not symbols.__contains__(character):
            tmp.append(character)
        elif character in "，。！？\n：；“”|）":
            tmp.append("。")
            for i in range(len(key_words)):
                if key_words[i] in ''.join(tmp):
                    sentences.append(''.join(tmp))
            tmp = []
        elif character == "\u3000":
            continue
    for i in range(len(key_words)):
        if key_words[i] in ''.join(tmp):
            sentences.append(''.join(tmp))
    return ''.join(sentences)

def filtrate_words(words,chinese_stopwords):
    find_chinese = re.compile(u"[\u4e00-\u9fa5]+")
    symbols = "[A-Za-z0-9\[\`\~\!\@\#\$\^\&\*\(\)\=\|\{\}\'\:\;\'\,\[\]\.\<\>\/\?\~\！\@\#\\\&\*\%]"        
    filtrated_words = []
    for j in range(len(words)):
        if re.findall(find_chinese,words[j]) == []:
            continue
        elif re.sub(symbols, "",re.findall(find_chinese,words[j])[0]) == '':
            continue
        elif re.sub(symbols, "",re.findall(find_chinese,words[j])[0]) in chinese_stopwords:
            continue
        else:
            filtrated_words.append(re.sub(symbols, "",re.findall(find_chinese,words[j])[0]))    
    return ' '.join(filtrated_words)

name = input('Please input the name of company: ')
current_path = os.getcwd()
if os.path.isdir('results/'+name+'/predict/results/') == False:
    os.makedirs(r'results/'+name+'/predict/results/')
path = 'data/'+name+'/'+name
model_load_path = 'results/'+name+'/train/model/'

chinese_stopwords = []
for line in open('data/stopwords.txt','rb'):
    chinese_stopwords.append(line.decode('utf-8-sig').split()[0])
key_words = []
for line in open(path+'_original_key_words.txt','rb'):
    key_words.append(line.decode('utf-8-sig').split()[0])
jieba.load_userdict(path+'_original_key_words.txt')

with open('data/'+name+'/preprocess/'+name+'_vocabulary_title.pkl','rb') as load1:
    vocabulary_title = pickle.load(load1)
with open('data/'+name+'/preprocess/'+name+'_idf_title.pkl','rb') as load2:
    idf_title = pickle.load(load2)
with open('results/'+name+'/train/model/'+'title_threshold.pkl','rb') as load3:
    Threshold = pickle.load(load3)
with open('results/'+name+'/train/model/'+'content_threshold.pkl','rb') as load4:
    threshold = pickle.load(load4)
with open('data/'+name+'/preprocess/'+name+'_vocabulary_content.pkl','rb') as load5:
    vocabulary_content = pickle.load(load5)
with open('data/'+name+'/preprocess/'+name+'_idf_content.pkl','rb') as load6:
    idf_content = pickle.load(load6)
 
workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = 'retain'
worksheet1.cell(row=1,column=1).value = 'title'
worksheet1.cell(row=1,column=2).value = 'content'
worksheet1.cell(row=1,column=3).value = 'label'
worksheet2 = workbook.create_sheet('delete')
worksheet2.cell(row=1,column=1).value = 'title'
worksheet2.cell(row=1,column=2).value = 'content'
worksheet2.cell(row=1,column=3).value = 'label'
count_retain = 2
count_delete = 2

excel = xlrd.open_workbook(path+'_test.xls')
table = excel.sheet_by_index(0)
num_rows = table.nrows-1

Labels = []
Predictions = []
for idx in range(1,num_rows):
    original_data = table.row_values(idx)
    label = int(original_data[2]=='保留')
    Labels.append(label)

    content = original_data[1]
    content_sentences = document2sentences(content,key_words)
    if content_sentences == '':
        prediction = 0
        Predictions.append(prediction)
        worksheet2.cell(row=count_delete,column=1).value = original_data[0].encode('gbk','ignore').decode('gbk','ignore')
        worksheet2.cell(row=count_delete,column=2).value = original_data[1].encode('gbk','ignore').decode('gbk','ignore')
        worksheet2.cell(row=count_delete,column=3).value = original_data[2].encode('gbk','ignore').decode('gbk','ignore')
        count_delete += 1
        print(name+' | Predict | Index | '+str(idx)+' | Delete')
    else:
        title = original_data[0]
        title_tokenized = jieba.lcut(title)
        content_sentences_tokenized = jieba.lcut(content_sentences)
        title_tokenized_filtered = filtrate_words(title_tokenized,chinese_stopwords)
        content_sentences_tokenized_filtered = filtrate_words(content_sentences_tokenized,chinese_stopwords)
        data_title = [5*(title_tokenized_filtered+' ')+content_sentences_tokenized_filtered]
        tf_transformer_title = CountVectorizer(ngram_range=(1,3),vocabulary=vocabulary_title)
        tf_title = tf_transformer_title.fit_transform(data_title)
        tf_weight_title = tf_title.toarray().tolist()
        tfidf_weight_title = normalize([[x*y for x,y in zip(tf_weight_title[0],idf_title)]], norm='l2').tolist()
        for ite in range(1,len(Threshold)+1):
            clf_title = joblib.load(model_load_path+name+'_iteration_'+str(ite)+'_train_title_classifier.m')
            tmp = clf_title.predict_proba(tfidf_weight_title).tolist()
            if tmp[0][1] < Threshold[ite]:
                prediction = 0
                Predictions.append(prediction)
                worksheet2.cell(row=count_delete,column=1).value = original_data[0].encode('gbk','ignore').decode('gbk','ignore')
                worksheet2.cell(row=count_delete,column=2).value = original_data[1].encode('gbk','ignore').decode('gbk','ignore')
                worksheet2.cell(row=count_delete,column=3).value = original_data[2].encode('gbk','ignore').decode('gbk','ignore')
                count_delete += 1
                print(name+' | Predict | Index | '+str(idx)+' | Delete')
                ite -= 1
                break
            else:
                continue
        if ite == len(Threshold):
            data_content = [content_sentences_tokenized_filtered]
            tf_transformer_content = CountVectorizer(ngram_range=(1,3),vocabulary=vocabulary_content)
            tf_content = tf_transformer_content.fit_transform(data_content)
            tf_weight_content = tf_content.toarray().tolist()
            tfidf_weight_content = normalize([[x*y for x,y in zip(tf_weight_content[0],idf_content)]], norm='l2').tolist()
            clf_content = joblib.load(model_load_path+name+'_train_content_classifier.m')
            tmp = clf_content.predict_proba(tfidf_weight_content).tolist()
            if tmp[0][1] < threshold:
                prediction = 0
                Predictions.append(prediction)
                worksheet2.cell(row=count_delete,column=1).value = original_data[0].encode('gbk','ignore').decode('gbk','ignore')
                worksheet2.cell(row=count_delete,column=2).value = original_data[1].encode('gbk','ignore').decode('gbk','ignore')
                worksheet2.cell(row=count_delete,column=3).value = original_data[2].encode('gbk','ignore').decode('gbk','ignore')
                count_delete += 1
                print(name+' | Predict | Index | '+str(idx)+' | Delete')
            else:
                prediction = 1
                Predictions.append(prediction)
                worksheet1.cell(row=count_retain,column=1).value = original_data[0].encode('gbk','ignore').decode('gbk','ignore')
                worksheet1.cell(row=count_retain,column=2).value = original_data[1].encode('gbk','ignore').decode('gbk','ignore')
                worksheet1.cell(row=count_retain,column=3).value = original_data[2].encode('gbk','ignore').decode('gbk','ignore')
                count_retain += 1
                print(name+' | Predict | Index | '+str(idx)+' | Retain')
workbook.save('results/'+name+'/predict/results/'+name+'_predict_results.xlsx')

print(name+' | Predict | Number of Data     | '+str(len(Labels)))
num_positive = Labels.count(1)
num_negative = Labels.count(0)
print(name+' | Predict | Number of Positive | '+str(num_positive))
print(name+' | Predict | Number of Negative | '+str(num_negative)+'\n')

recall = metrics.recall_score(Labels,Predictions,pos_label=1)
precision = metrics.precision_score(Labels,Predictions,pos_label=1)
f1 = metrics.f1_score(Labels,Predictions,pos_label=1) 

print(name+' | Predict | Positive Recall    | ' + '%.4f'%recall)
print(name+' | Predict | Positive Precision | ' + '%.4f'%precision)
print(name+' | Predict | Positive F1        | ' + '%.4f'%f1+'\n')

